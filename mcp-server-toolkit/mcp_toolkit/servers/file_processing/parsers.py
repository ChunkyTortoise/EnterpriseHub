"""File type detection and parsing for PDF, CSV, Excel, TXT, JSON, and Markdown."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class ParsedFile:
    """Result of parsing a file."""

    filename: str
    file_type: str
    text: str
    metadata: dict[str, Any] = field(default_factory=dict)
    pages: int = 1
    rows: int = 0
    error: str | None = None

    @property
    def is_success(self) -> bool:
        return self.error is None


class FileParser:
    """Detects file type and extracts text content.

    Supports: PDF, CSV, Excel (.xlsx), TXT, JSON, Markdown.
    """

    SUPPORTED_EXTENSIONS = {
        ".pdf": "pdf",
        ".csv": "csv",
        ".xlsx": "excel",
        ".xls": "excel",
        ".txt": "text",
        ".json": "json",
        ".md": "markdown",
        ".markdown": "markdown",
    }

    def detect_type(self, filename: str) -> str:
        """Detect file type from extension."""
        ext = Path(filename).suffix.lower()
        return self.SUPPORTED_EXTENSIONS.get(ext, "unknown")

    async def parse(self, content: bytes, filename: str) -> ParsedFile:
        """Parse file content based on detected type."""
        file_type = self.detect_type(filename)

        parsers = {
            "pdf": self._parse_pdf,
            "csv": self._parse_csv,
            "excel": self._parse_excel,
            "text": self._parse_text,
            "json": self._parse_json,
            "markdown": self._parse_text,
        }

        parser = parsers.get(file_type)
        if parser is None:
            return ParsedFile(
                filename=filename,
                file_type="unknown",
                text="",
                error=f"Unsupported file type: {Path(filename).suffix}",
            )

        try:
            return await parser(content, filename)
        except Exception as e:
            return ParsedFile(
                filename=filename,
                file_type=file_type,
                text="",
                error=f"Parse error: {e}",
            )

    async def _parse_pdf(self, content: bytes, filename: str) -> ParsedFile:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return ParsedFile(
                filename=filename,
                file_type="pdf",
                text="",
                error="PyPDF2 not installed. Install with: pip install PyPDF2",
            )

        reader = PdfReader(io.BytesIO(content))
        pages = []
        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text)

        return ParsedFile(
            filename=filename,
            file_type="pdf",
            text="\n\n".join(pages),
            pages=len(reader.pages),
            metadata={
                "page_count": len(reader.pages),
                "info": {k: str(v) for k, v in (reader.metadata or {}).items()},
            },
        )

    async def _parse_csv(self, content: bytes, filename: str) -> ParsedFile:
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        headers = reader.fieldnames or []

        formatted_lines = [", ".join(headers)]
        for row in rows:
            formatted_lines.append(", ".join(str(row.get(h, "")) for h in headers))

        return ParsedFile(
            filename=filename,
            file_type="csv",
            text="\n".join(formatted_lines),
            rows=len(rows),
            metadata={"headers": headers, "row_count": len(rows)},
        )

    async def _parse_excel(self, content: bytes, filename: str) -> ParsedFile:
        try:
            import openpyxl
        except ImportError:
            return ParsedFile(
                filename=filename,
                file_type="excel",
                text="",
                error="openpyxl not installed. Install with: pip install openpyxl",
            )

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        all_text = []
        total_rows = 0
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            all_text.append(f"## Sheet: {sheet_name}")
            rows = list(ws.iter_rows(values_only=True))
            total_rows += len(rows)
            for row in rows:
                all_text.append(", ".join(str(c) if c is not None else "" for c in row))

        wb.close()
        return ParsedFile(
            filename=filename,
            file_type="excel",
            text="\n".join(all_text),
            rows=total_rows,
            metadata={"sheets": sheet_names, "total_rows": total_rows},
        )

    async def _parse_text(self, content: bytes, filename: str) -> ParsedFile:
        text = content.decode("utf-8", errors="replace")
        line_count = text.count("\n") + 1
        return ParsedFile(
            filename=filename,
            file_type=self.detect_type(filename),
            text=text,
            metadata={"line_count": line_count, "char_count": len(text)},
        )

    async def _parse_json(self, content: bytes, filename: str) -> ParsedFile:
        text = content.decode("utf-8", errors="replace")
        data = json.loads(text)
        formatted = json.dumps(data, indent=2)
        row_count = len(data) if isinstance(data, list) else 0
        return ParsedFile(
            filename=filename,
            file_type="json",
            text=formatted,
            rows=row_count,
            metadata={
                "type": type(data).__name__,
                "keys": list(data.keys()) if isinstance(data, dict) else [],
                "item_count": len(data) if isinstance(data, (list, dict)) else 0,
            },
        )
